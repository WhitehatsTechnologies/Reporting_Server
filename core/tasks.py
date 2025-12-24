# =============== Built-in Modules ===============
import os
import re
import csv
import json
import tempfile
import traceback
from datetime import datetime
from pprint import pprint
import time
# ================================================

# =============== Third-party Modules ===============
import pandas as pd
from openpyxl import Workbook
from django.conf import settings
from celery import shared_task
# ===================================================

# =============== DataForesight Modules ===============
from common.custom_utils import log_traceback
from common.common_utils import update_celery_config
# =====================================================

# =============== Local Modules ===============
from Reporting_Server.settings import CONNECTOR_TYPE
# =============================================

_ILLEGAL_XML_RE = re.compile(r'[\x00-\x08\x0B\x0C\x0E-\x1F]')
MAX_EXCEL_CELL_CHARS = 32767

def get_chunked_ids(queryset, chunk_size=100_000):
    """
    Yields lists of PKs in chunks, ordered by PK.
    """
    last_pk = 0
    qs = queryset.order_by("pk").values_list("pk", flat=True)
    while True:
        chunked_ids = list(qs.filter(pk__gt=last_pk)[:chunk_size])
        if not chunked_ids:
            break
        last_pk = chunked_ids[-1]
        yield chunked_ids


def _clean_string(s):
    # remove illegal XML chars
    s = _ILLEGAL_XML_RE.sub("", s)
    # optionally trim trailing nulls/whitespace
    if len(s) > MAX_EXCEL_CELL_CHARS:
        return s[:MAX_EXCEL_CELL_CHARS]
    return s


def _clean_cell_value(val):
    from datetime import datetime, date, time, timedelta

    # Keep common native types intact
    if val is None:
        return None
    if isinstance(val, (int, float, bool)):
        return val
    if isinstance(val, (datetime, date, time, timedelta)):
        return val
    # If it's a list/tuple/dict — stringify safely
    if isinstance(val, (list, tuple, dict)):
        try:
            s = json.dumps(val, default=str, ensure_ascii=False)
        except Exception:
            s = str(val)
        return _clean_string(s)
    # Fallback: convert to string and clean
    try:
        s = str(val)
    except Exception:
        s = repr(val)
    return _clean_string(s)


def merge_report_chunks(customer_db, report_id, report_chunks, merged_report_path, report_format="csv", sheet_title='Sheet'):
    """
    Loads and merges pickled DataFrame chunks in bulk per file, avoiding row-by-row writes.
    """
    report_obj_up = GenReport.objects.using(customer_db).filter(id=report_id)
    report_obj = report_obj_up.first()
    if report_obj:
        report_format = report_obj.report_format.lower()    
    report_progress = 0
    progress_per_chunk = 100 / len(report_chunks) if report_chunks else 0

    # CSV merging: write header and bulk rows per chunk
    if report_format.lower() == 'csv':
        with open(merged_report_path, 'w', newline='', encoding='utf-8') as out_file:
            writer = csv.writer(out_file)
            header_written = False
            for path in report_chunks:
                try:
                    df = pd.read_pickle(path)
                    print(f'Loaded chunk {path} with {len(df)} rows')
                    if not header_written:
                        writer.writerow(df.columns.tolist())
                        header_written = True
                    # bulk write all rows from this chunk
                    writer.writerows(df.itertuples(index=False, name=None))
                    del df

                    report_progress += progress_per_chunk
                    report_progress = round(report_progress, 2)
                    report_obj_up.update(report_progress=report_progress)
                    print(f'merging progress: {report_progress}%')
                        
                except Exception as e:
                    print(f"Error processing {path}: {e}")

    # # Excel merging: use pandas ExcelWriter to bulk write each chunk
    # elif report_format.lower() == 'excel':
    #     # Excel limits: 1,048,576 rows per sheet
    #     max_rows = 1000000  # 1 million rows per sheet
    #     wb = Workbook(write_only=True)
    #     sheet_index = 1
    #     ws = wb.create_sheet(title=f"{sheet_title} {sheet_index}")
    #     header_written = False
    #     current_row_count = 0

    #     for path in report_chunks:
    #         # try:
    #         df = pd.read_pickle(path)
    #         rows = df.values.tolist()
    #         print(f'Loaded chunk {path} with {len(rows)} rows')

    #         # if adding this chunk would overflow, start a new sheet
    #         if header_written and current_row_count + len(rows) > max_rows:
    #             sheet_index += 1
    #             ws = wb.create_sheet(title=f"{sheet_title} {sheet_index}")
    #             header_written = False
    #             current_row_count = 0

    #         # write header at top of each new sheet
    #         if not header_written:
    #             ws.append(df.columns.tolist())
    #             header_written = True
    #             current_row_count += 1

    #         # bulk append chunk rows
    #         for row in rows:
    #             ws.append(row)
    #         current_row_count += len(rows)
    #         del df

    #         report_progress += progress_per_chunk
    #         report_progress = round(report_progress, 2)
    #         report_obj_up.update(report_progress=report_progress) 
    #         print(f'merging progress: {report_progress}%')

    #         # except Exception as e:
    #         #     print(f"Error processing {path}: {e}")

    #     wb.save(merged_report_path)

    # Excel merging: use write_only Workbook and bulk append cleaned rows
    elif report_format.lower() == 'excel':
        # Excel limits: ~1,048,576 rows per sheet; we use 1,000,000 for safety
        max_rows = 1000000
        wb = Workbook(write_only=True)
        sheet_index = 1
        ws = wb.create_sheet(title=f"{sheet_title} {sheet_index}")
        header_written = False
        current_row_count = 0
        try:

            for path in report_chunks:
                try:
                    df = pd.read_pickle(path)
                except Exception as e:
                    print(f"Error loading {path}: {e}")
                    continue

                # Convert DataFrame chunk to list-of-lists and clean each cell
                # Use df.itertuples for memory efficiency if needed; here we do values
                rows = df.values.tolist()
                print(f'Loaded chunk {path} with {len(rows)} rows')

                # if adding this chunk would overflow, start a new sheet
                if header_written and current_row_count + len(rows) > max_rows:
                    sheet_index += 1
                    ws = wb.create_sheet(title=f"{sheet_title} {sheet_index}")
                    header_written = False
                    current_row_count = 0

                # write header at top of each new sheet (clean header strings)
                if not header_written:
                    cleaned_header = [ _clean_cell_value(c) for c in df.columns.tolist() ]
                    ws.append(cleaned_header)
                    header_written = True
                    current_row_count += 1

                # bulk append cleaned chunk rows
                for r in rows:
                    cleaned_row = [ _clean_cell_value(cell) for cell in r ]
                    ws.append(cleaned_row)

                current_row_count += len(rows)
                del df

                report_progress += progress_per_chunk
                report_progress = round(report_progress, 2)
                report_obj_up.update(report_progress=report_progress)
                print(f'merging progress: {report_progress}%')
        except Exception as e:
            print(f"Error processing {path}: {e}")

        # save workbook
        try:
            wb.save(merged_report_path)
        except Exception as e:
            # If save fails, print the exception for debugging
            print(f"Error saving workbook to {merged_report_path}: {e}")
        

    else:
        raise ValueError(f"Unsupported report format: {report_format}.")
    
    for path in report_chunks:
        try:
            os.remove(path)
        except Exception:
            print(traceback.format_exc())

    print(f'Merged report written to: {merged_report_path}')
    return merged_report_path



@shared_task
def write_column_report_chunk(customer_db, report_id, chunked_ids, tmp_dir):
    """
    Write a chunk of ColumnName data to a temporary pickle file.
    """
    update_celery_config(customer_dbs_to_update=[customer_db])
    
    # Building the temporary chunked report path.
    report_chunk = os.path.join(
        tmp_dir,
        f"chunked_column_report_{customer_db}_{report_id}_{chunked_ids[0]}-{chunked_ids[-1]}.pkl"
    )
    print(f'report_chunk: {report_chunk}')

    chunked_column_name_objs = ColumnName.objects.using(customer_db).filter(pk__in=chunked_ids)
    
    # Fetching rows in a list.
    column_names_list = list(
        chunked_column_name_objs
        .values(
            "id",
            "column_name",
            "table_name_obj__table_name",
            "table_name_obj__db_name_obj__db_name",
            "table_name_obj__db_name_obj__host_obj__id",
            "table_name_obj__db_name_obj__host_obj__scan_host",
            "table_name_obj__db_name_obj__host_obj__scan_platform",
            "scan_status",
            "cell_count",
        )
    )

    sb_totals = (
        ScanBreakdown.objects.using(customer_db)
        .filter(scan_unit_obj__column_name_obj__in=chunked_ids, data_count__gt=0)
        .values('scan_unit_obj__column_name_obj__id')
        .annotate(total=Sum('data_count'))
    )
    totals_by_col = {d['scan_unit_obj__column_name_obj__id']: d['total'] for d in sb_totals}

    host_ids = {d['table_name_obj__db_name_obj__host_obj__id'] for d in column_names_list}
    hosts = Host.objects.using(customer_db).filter(id__in=host_ids).prefetch_related('host_tags')
    host_map = {h.id: h for h in hosts}

    for column_name_dict in column_names_list:
        column_id = column_name_dict["id"]
        column_name_dict['data_found'] = totals_by_col.get(column_id, 0)
        
        host_id = column_name_dict["table_name_obj__db_name_obj__host_obj__id"]
        host_obj = host_map[host_id]
        host_tags_str = ', '.join(t.tag_name for t in host_obj.host_tags.all()) or 'N/A'
        column_name_dict["host_tags"] = host_tags_str

        if not column_name_dict["scan_status"]:
            column_name_dict["scan_status"] = "Not Scanned Yet"
        
        # Removing some keys which is not required in final DataFrame.
        del column_name_dict["id"] 
        del column_name_dict["table_name_obj__db_name_obj__host_obj__id"]

    print(f'Creating DataFrame with {len(column_names_list)} rows')
    df = pd.DataFrame(column_names_list)
    print(f'DataFrame created with shape: {df.shape}')

    # 4) Rename columns to your desired field names
    df.rename(
        columns={
            "column_name": "Column",
            "table_name_obj__table_name": "Table",
            "table_name_obj__db_name_obj__db_name": "Database",
            "table_name_obj__db_name_obj__host_obj__scan_host": "Host",
            "table_name_obj__db_name_obj__host_obj__scan_platform": "Platform",
            "host_tags": "Host Tags",
            "scan_status": "Scan Status",
            "data_found": "Data Found",
            "cell_count": "Cell Count",
        },
        inplace=True
    )
    print(f'Renaming columns completed. DataFrame shape: {df.shape}')

    desired_order = [
        "Column", "Table", "Database",
        "Host", "Platform", "Host Tags",
        "Scan Status", "Data Found", "Cell Count"
    ]

    # This returns a new DataFrame with exactly that column sequence:
    df = df[desired_order]

    # serialize to pickle
    df.to_pickle(report_chunk)

    print(f'Chunked pickle written to: {report_chunk}')
    return report_chunk


@shared_task
def write_host_details_report_chunk(customer_db, report_id, chunked_ids, tmp_dir):
    """
    Write a chunk of Host data to a temporary pickle file for the Host Details report.
    Outputs columns: Scan Host, Platform Type, Platform, Service Account, Host Tags, Departments, Details
    """
    update_celery_config(customer_dbs_to_update=[customer_db])

    report_chunk = os.path.join(
        tmp_dir,
        f"chunked_host_details_report_{customer_db}_{report_id}_{chunked_ids[0]}-{chunked_ids[-1]}.pkl"
    )
    print(f'report_chunk: {report_chunk}')

    # Fetch Host objects and prefetch M2M relations for tags and departments
    host_qs = Host.objects.using(customer_db).filter(pk__in=chunked_ids).prefetch_related('host_tags', 'department_objs')

    rows = []
    for host in host_qs:
        try:
            scan_host = host.scan_host
            platform_type = (host.platform_type or '').strip()
            scan_platform = host.scan_platform
            scanner_type = host.scanner_type
            location = host.location_type

            # Service account display fallback
            svc_display = 'N/A'
            try:
                if host.service_account:
                    svc = host.service_account
                    svc_display = getattr(svc, 'service_account_name', None) or getattr(svc, 'username', None) or str(getattr(svc, 'id', 'N/A'))
            except Exception:
                svc_display = 'N/A'

            # Host tags
            try:
                host_tags = ', '.join(t.tag_name for t in host.host_tags.all()) or 'N/A'
            except Exception:
                host_tags = 'N/A'

            # Departments
            try:
                departments = ', '.join(d.department_name for d in host.department_objs.all()) or 'N/A'
            except Exception:
                departments = 'N/A'

            # Platform-specific details
            pt = (platform_type or '').lower().strip()
            details = ''
            if pt in DRIVE_PLATFORMS:
                details = f"IP: {host.ip_address or 'N/A'}; Hostname: {host.hostname or 'N/A'}; Shared Drive: {host.shared_drive_name or 'N/A'}"
            elif pt in DB_PLATFORMS:
                details = f"DB Instance: {host.db_instance or 'N/A'}; Port: {host.db_port or 'N/A'}; DB Account: {host.db_account or 'N/A'}"
            elif pt in CLOUD_PLATFORMS:
                details = f"Cloud Host: {scan_host or 'N/A'}; Client ID present: {'Yes' if host.client_id else 'No'}"
            elif pt == 'email':
                details = f"Email: {host.email_address or 'N/A'}"
            else:
                # Generic fallback
                details = f"IP: {host.ip_address or 'N/A'}; Hostname: {host.hostname or 'N/A'}"
                
            # Show Service Account only for drive and db platform types (case-insensitive match)
            show_service_account = False
            if platform_type.lower() == 'db' or platform_type.lower() == 'drive':
                show_service_account = True

            row = {
                'Scan Host': scan_host,
                'Platform Type': platform_type,
                'Scan Platform': scan_platform,
                'Location': location,
                'Scanner Type': scanner_type,
                'Host Tags': host_tags,
                'Departments': departments,
                'Details': details,
            }

            if show_service_account:
                row['Service Account'] = svc_display

            rows.append(row)

        except Exception:
            # Skip problematic host but log for debugging
            print(traceback.format_exc())
            continue

    print(f'Creating DataFrame with {len(rows)} rows')
    df = pd.DataFrame(rows)
    print(f'DataFrame created with shape: {df.shape}')

    if platform_type.lower() == 'drive' or platform_type.lower() == 'db':

        desired_order = [
            'Scan Host', 'Platform Type', 'Scan Platform', 'Location', 'Scanner Type', 'Service Account', 'Host Tags', 'Departments', 'Details'
        ]
    else:
        desired_order = [
            'Scan Host', 'Platform Type', 'Scan Platform', 'Location', 'Scanner Type', 'Host Tags', 'Departments', 'Details'
        ]

    # Ensure columns exist
    for col in desired_order:
        if col not in df.columns:
            df[col] = 'N/A'

    df = df[desired_order]

    # serialize to pickle
    df.to_pickle(report_chunk)

    print(f'Chunked host details pickle written to: {report_chunk}')
    return report_chunk


@shared_task
def write_audit_log_report_chunk(customer_db, report_id, chunked_ids, tmp_dir):
    """
    Write a chunk of AuditLog data to a temporary pickle file for the Audit Log report.
    Columns: User, Activity, Details, Panel, Is Notification, Is Notified, Notification Link, Created At
    """
    update_celery_config(customer_dbs_to_update=[customer_db])

    report_chunk = os.path.join(
        tmp_dir,
        f"chunked_audit_log_report_{customer_db}_{report_id}_{chunked_ids[0]}-{chunked_ids[-1]}.pkl"
    )
    print(f'report_chunk: {report_chunk}')

    # Query AuditLog rows for this chunk
    audit_qs = None
    try:
        # If a queryset of AuditLog is available in DB, filter by ids
        audit_qs = AuditLog.objects.using(customer_db).filter(pk__in=chunked_ids).select_related('user')
    except Exception:
        print('AuditLog model not found in expected locations')
        audit_qs = []

    rows = []
    for a in audit_qs:
        try:
            user_display = 'Anonymous'
            try:
                if a.user:
                    user_display = getattr(a.user, 'username', None) or getattr(a.user, 'email', None) or str(a.user.id)
            except Exception:
                user_display = 'Anonymous'

            panel = 'User' if a.user_panel else ('Admin Panel' if a.admin_panel else 'Other')

            rows.append(
                {
                    'User': user_display,
                    'Log Source': panel,
                    'Activity': a.activity,
                    'Details': a.details or '',
                    'Created At': a.created_at,
                }
            )
        except Exception:
            print(traceback.format_exc())
            continue

    print(f'Creating DataFrame with {len(rows)} rows')
    df = pd.DataFrame(rows)
    print(f'DataFrame created with shape: {df.shape}')

    # Format Created At
    if 'Created At' in df.columns:
        try:
            df['Created At'] = pd.to_datetime(df['Created At']).dt.strftime('%Y-%m-%d %H:%M:%S')
        except Exception:
            pass

    desired_order = [
        'User', 'Log Source', 'Activity', 'Details', 'Created At'
    ]

    for col in desired_order:
        if col not in df.columns:
            df[col] = ''

    df = df[desired_order]

    df.to_pickle(report_chunk)
    print(f'Chunked audit log pickle written to: {report_chunk}')
    return report_chunk



# Unused
def generate_column_report_old(customer_db: str, report_id: int):
    """
    Generates report of database columns.
    """

    column_report_obj_up = ColumnReport.objects.using(customer_db).filter(id=int(report_id))
    column_report_obj = column_report_obj_up.first()
    report_id = column_report_obj.id
    filter_inputs = column_report_obj.filter_inputs
    cell_gte = column_report_obj.cell_gte
    report_format = column_report_obj.report_type
    worker_limit = column_report_obj.worker_limit

    extension = "xlsx" if report_format.lower() == "excel" else "csv"
    report_name = f"column_report_{report_id}.{extension}"
    report_path = os.path.join(settings.MEDIA_ROOT, report_name)

    # Getting connector object.
    connector_obj = get_connector_obj(
        connector_type=CONNECTOR_TYPE,
        private_ip=PRIVATE_IP,
        customer_db=customer_db,
    )

    column_report_obj_up.update(
        connector_obj=connector_obj,
        report_name=report_name, 
        report_status="Processing"
    )

    # Getting column name objects.
    column_name_objs = filter_column_name_objs(customer_db, filter_inputs)
    column_name_objs = column_name_objs.filter(cell_count__gte=cell_gte) if cell_gte else column_name_objs

    total_column_names = column_name_objs.count()
    chunk_size = settings.REPORT_CHUNK_SIZE
    progress_per_chunk = 100 / (total_column_names // chunk_size + 1) if total_column_names > chunk_size else 100

    column_report_obj_up.update(report_status="Writing")

    assigned_tasks = []
    report_chunks = []
    progress = 0
    for chunked_ids in get_chunked_ids(column_name_objs, chunk_size=chunk_size):
        task = write_column_report_chunk.delay(
            customer_db=customer_db,
            report_id=report_id,
            chunked_ids=chunked_ids,
            tmp_dir=settings.MEDIA_ROOT,
        )
        assigned_tasks.append(task)

        while len(assigned_tasks) >= worker_limit:
            for task in assigned_tasks.copy():
                try:
                    if task.ready():
                        report_chunks.append(task.get())
                        assigned_tasks.remove(task)
                        progress += progress_per_chunk
                        print(f"writing progress: {progress}%")
                except Exception as e:
                    print(f"Error processing task {task.id}: {e}")
                    assigned_tasks.remove(task)
                    progress += progress_per_chunk
            
            report_progress = round(progress, 2)
            column_report_obj_up.update(report_progress=report_progress)
            time.sleep(1)

    # Handle any remaining tasks
    while len(assigned_tasks) > 0:
        for task in assigned_tasks.copy():
            try:
                if task.ready():
                    report_chunks.append(task.get())
                    assigned_tasks.remove(task)
                    progress += progress_per_chunk
                    print(f"writing progress: {progress}%")
            except Exception as e:
                print(f"Error processing task {task.id}: {e}")
                assigned_tasks.remove(task)
                progress += progress_per_chunk

        report_progress = round(progress, 2)
        column_report_obj_up.update(report_progress=report_progress)
        time.sleep(1)

    
    column_report_obj_up.update(report_status="Merging")

    # Merging all chunked reports into a single file.
    merge_report_chunks(
        customer_db=customer_db,
        report_id=report_id,
        report_chunks=report_chunks,
        merged_report_path=report_path,
        report_format=report_format,
        sheet_title=f"Column Report"
    )

    # Final updating the report object.
    column_report_obj_up.update(
        report_path=report_path,
        report_status="Generated",
    )

    print(f"Column report generated: {report_path}")
    return



@shared_task
def write_scan_report_chunk(customer_db, report_id, chunked_ids, tmp_dir):
    """
    Write a chunk of Scan data to a temporary pickle file.
    """
    update_celery_config(customer_dbs_to_update=[customer_db])
    
    # Building the temporary chunked report path.
    report_chunk = os.path.join(
        tmp_dir,
        f"scan_report_chunk_{customer_db}_{report_id}_{chunked_ids[0]}-{chunked_ids[-1]}.pkl"
    )
    print(f'report_chunk: {report_chunk}')

    report_obj = GenReport.objects.using(customer_db).get(id=report_id)
    mask_status = report_obj.mask_status

    if mask_status == 'masked':
        data_field = 'data_obj__data'
    elif mask_status == 'unmasked':
        data_field = 'data_obj__value_obj__encrypted_value'
    else:
        raise ValueError(f"Invalid mask status: {mask_status}")

    fields = [
        "id",
        "scan_obj__id",
        "host_obj__scan_host",
        "host_obj__scan_platform",
        "host_obj__platform_type",
        "scan_unit_obj__file_path",
        "scan_unit_obj__db_name_obj__db_name",
        "scan_unit_obj__table_name_obj__table_name",
        "scan_unit_obj__column_name_obj__column_name",
        "scan_unit_obj__email_subject",
        "scan_unit_obj__email_label",
        data_field,
        "data_type_obj__data_type",
    ]

    chunked_result_objs = (
        Result.objects.using(customer_db)
        .filter(pk__in=chunked_ids)
        .values(*fields)
        .distinct()
    )
    
    # Fetching rows in a list.
    result_objs_list = list(chunked_result_objs)

    for result_obj in result_objs_list:
        if mask_status == "unmasked":
            encrypted_data = result_obj["data_obj__value_obj__encrypted_value"]
            decrypted_data = decrypt_data(encrypted_data)
            result_obj["data_obj__data"] = decrypted_data
            del result_obj["data_obj__value_obj__encrypted_value"]
        
        platform_type_lower = result_obj.get("host_obj__platform_type", "").lower().strip()

        if platform_type_lower == 'drive' or platform_type_lower == 'cloud':
            data_location = f"File path: {result_obj['scan_unit_obj__file_path']}"
            del result_obj["scan_unit_obj__file_path"]

        elif platform_type_lower == 'db':
            data_location = f"Database: {result_obj['scan_unit_obj__db_name_obj__db_name']}, Table: {result_obj['scan_unit_obj__table_name_obj__table_name']}, Column: {result_obj['scan_unit_obj__column_name_obj__column_name']}"
            del result_obj["scan_unit_obj__db_name_obj__db_name"]
            del result_obj["scan_unit_obj__table_name_obj__table_name"]
            del result_obj["scan_unit_obj__column_name_obj__column_name"]
        
        elif platform_type_lower == 'email':
            data_location = f"Labels: {result_obj['scan_unit_obj__email_label']}, Email Subject: {result_obj['scan_unit_obj__email_subject']} "
            del result_obj["scan_unit_obj__email_subject"]
            del result_obj["scan_unit_obj__email_label"]

        result_obj["data_location"] = data_location


    print(f'Creating DataFrame with {len(result_objs_list)} rows')
    df = pd.DataFrame(result_objs_list)
    df = df.drop_duplicates(subset=['id'], keep='first').reset_index(drop=True) # Ensuring unique IDs.
    print(f'DataFrame created with shape: {df.shape}')

    # Renaming columns to desired field names.
    df.rename(
        columns={
            "scan_obj__id": "Scan ID",
            "host_obj__scan_host": "Host",
            "host_obj__scan_platform": "Platform",
            "data_location": "Data Location",
            "data_obj__data": "Data",
            "data_type_obj__data_type": "Data Type",
        },
        inplace=True
    )
    print(f'Renaming columns completed. DataFrame shape: {df.shape}')

    desired_order = [
        "Scan ID", "Host", "Platform",
        "Data Location", "Data", "Data Type"
    ]

    # This returns a new DataFrame with exactly that column sequence:
    df = df[desired_order]

    # serialize to pickle
    df.to_pickle(report_chunk)

    print(f'Chunked pickle written to: {report_chunk}')
    return report_chunk


@shared_task
def write_scan_status_report_chunk(customer_db, report_id, chunked_ids, tmp_dir):
    """
    Build a per-scan chunk:
      - Rows come from Scan (chunked_ids are Scan IDs)
      - Per–Data-Type counts come from ScanBreakdown grouped by Scan ID
      - ALL scans are preserved. If no breakdown rows exist for a Scan ID:
          * Total = 0
          * Data-type columns = 0 (and we add a placeholder 'N/A' column if none exist)
      - Output columns:
        ["Scan ID","Host","Platform","Scanner Type","Scan Start Time","Scan Status",
         "Scanned Content Size","Scan Details", <Data Type columns...>, "Total"]
    """
    # Best-effort celery config update
    try:
        update_celery_config(customer_dbs_to_update=[customer_db])
    except Exception:
        pass

    # Chunk path
    start_id = chunked_ids[0] if chunked_ids else "none"
    end_id = chunked_ids[-1] if chunked_ids else "none"
    fname = f"scan_status_report_chunk_{customer_db}_{report_id}_{start_id}-{end_id}.pkl"
    report_chunk_path = os.path.join(tmp_dir, fname)
    os.makedirs(tmp_dir, exist_ok=True)
    print(f"[write_scan_status_report_chunk] writing: {report_chunk_path} (scan_count={len(chunked_ids)})")

    # 1) Fetch Scan rows (row source)
    scans = list(
        Scan.objects.using(customer_db)
        .filter(pk__in=chunked_ids)
        .values(
            "id",
            "host_obj__scan_host",
            "host_obj__scan_platform",
            "host_obj__scanner_type",
            "scan_status",
            "scan_start_time",
            "scanned_content_size",
            "scan_target",
            "data_types",
            "scan_extensions",
            "search_extensions",
            "scan_dbs",
            "scan_tables",
            "scan_columns",
            "scan_labels",
            "scan_sections",
            "scan_paths",
        )
    )

    if not scans:
        empty_cols = [
            "Scan ID","Host","Platform","Scanner Type","Scan Start Time","Scan Status",
            "Scanned Content Size","Scan Details","N/A","Total"
        ]
        df_empty = pd.DataFrame(columns=empty_cols)
        df_empty.to_pickle(report_chunk_path)
        print(f"[write_scan_status_report_chunk] empty chunk written: {report_chunk_path}")
        return report_chunk_path

    df_scan = pd.DataFrame(scans).rename(columns={
        "id": "Scan ID",
        "host_obj__scan_host": "Host",
        "host_obj__scan_platform": "Platform",
        "host_obj__scanner_type": "Scanner Type",
        "scan_status": "Scan Status",
        "scan_start_time": "Scan Start Time",
        "scanned_content_size": "Scanned Content Size (in KB)",
        "scan_target": "Scan Target",
        "data_types": "Data Types",
        "scan_extensions": "Scan Extensions",
        "search_extensions": "Search Extensions",
        "scan_dbs": "Scan DBs",
        "scan_tables": "Scan Tables",
        "scan_columns": "Scan Columns",
        "scan_labels": "Scan Labels",
        "scan_sections": "Scan Sections",
    })

    # Basic cleaning
    df_scan["Scan ID"] = df_scan["Scan ID"].astype(str)
    for t in ["Host","Platform","Scanner Type","Scan Status"]:
        df_scan[t] = df_scan[t].fillna("").astype(str).str.strip()

    # Convert scanned content size to KB
    df_scan["Scanned Content Size (in KB)"] = (
        pd.to_numeric(df_scan["Scanned Content Size (in KB)"], errors="coerce")
        .fillna(0)
        .astype(int) // 1024
    )

    df_scan["Scan Start Time"] = pd.to_datetime(df_scan["Scan Start Time"], errors="coerce").dt.strftime("%Y-%m-%d %H:%M:%S")
    df_scan["Scan Start Time"] = df_scan["Scan Start Time"].fillna("")

    # Scan Details builder
    def split_clean(s):
        if s is None:
            return []
        parts = [p.strip() for p in re.split(r"[,\|;]+", str(s)) if p is not None]
        return [p for p in parts if p and p.lower() not in ("none", "nan")]

    def make_details(row):
        pieces = []
        def add(label, key):
            vals = split_clean(row.get(key, ""))
            if vals:
                pieces.append(f"{label}: " + ", ".join(vals))
        add("Scan Target", "Scan Target")
        add("Data Types", "Data Types")
        add("Scan Extensions", "Scan Extensions")
        add("Search Extensions", "Search Extensions")
        add("Scan DBs", "Scan DBs")
        add("Scan Tables", "Scan Tables")
        add("Scan Columns", "Scan Columns")
        add("Scan Labels", "Scan Labels")
        add("Scan Sections", "Scan Sections")
        add("Scan Paths", "scan_paths")
        add("File Paths", "file_paths")
        add("File Paths To Scan", "file_paths_to_scan")
        return " | ".join(pieces) if pieces else ""

    df_scan["Scan Details"] = df_scan.apply(make_details, axis=1)

    # Keep int Scan IDs for breakdown query
    scan_ids_int = []
    for x in df_scan["Scan ID"]:
        try:
            scan_ids_int.append(int(x))
        except Exception:
            pass

    # 2) Fetch per–Data-Type counts from ScanBreakdown
    scan_breakdown_rows = list(
        ScanBreakdown.objects.using(customer_db)
        .filter(scan_obj_id__in=scan_ids_int)
        .values("scan_obj_id", "data_type_obj__data_type")
        .annotate(data_count_sum=Sum("data_count"))
    )

    if scan_breakdown_rows:
        scan_breakdown_df = pd.DataFrame(scan_breakdown_rows).rename(columns={
            "scan_obj_id": "Scan ID",
            "data_type_obj__data_type": "Data Type",
            "data_count_sum": "Data Count",
        })
        scan_breakdown_df["Scan ID"] = scan_breakdown_df["Scan ID"].astype(str)
        scan_breakdown_df["Data Count"] = pd.to_numeric(scan_breakdown_df["Data Count"], errors="coerce").fillna(0).astype(int)

        # Pivot to one column per Data Type
        df_pivot = scan_breakdown_df.pivot_table(
            index="Scan ID",
            columns="Data Type",
            values="Data Count",
            fill_value=0,
            aggfunc="sum",
        ).reset_index()
        df_pivot.columns.name = None
    else:
        # No breakdown rows at all → just a frame with Scan ID
        df_pivot = df_scan[["Scan ID"]].copy()

    # 3) Merge (LEFT) so ALL scans are kept
    meta_cols = ["Scan ID","Host","Platform","Scanner Type","Scan Start Time","Scan Status","Scanned Content Size (in KB)","Scan Details"]
    df_meta = df_scan[meta_cols].drop_duplicates(subset=["Scan ID"])
    result = df_meta.merge(df_pivot, on="Scan ID", how="left")

    # Identify data-type columns and normalize
    fixed = set(meta_cols + ["Total"])
    dtype_cols = [c for c in result.columns if c not in fixed and c != "Scan ID"]

    # If no dtype columns exist (no breakdown in this chunk), add placeholder 'N/A'
    if not dtype_cols:
        result["N/A"] = 0
        dtype_cols = ["N/A"]

    # Ensure dtype cols numeric and fill NaNs with 0
    for c in dtype_cols:
        result[c] = pd.to_numeric(result[c], errors="coerce").fillna(0).astype(int)

    # Total = sum of dtype columns
    result["Total"] = result[dtype_cols].sum(axis=1).astype(int)

    # Final column order
    final_cols = meta_cols + sorted(dtype_cols) + ["Total"]
    result = result[final_cols]

    # 4) Persist chunk
    result.to_pickle(report_chunk_path)
    print(f"[write_scan_status_report_chunk] chunk written: {report_chunk_path} rows={len(result)} dtype_cols={len(dtype_cols)}")
    return report_chunk_path


# unused
def generate_scan_report_old(customer_db: str, report_id: int):
    """
    Generates report of scan results.
    """

    scan_report_obj_up = GenReport.objects.using(customer_db).filter(id=int(report_id))
    scan_report_obj = scan_report_obj_up.first()
    report_id = scan_report_obj.id
    report_format = scan_report_obj.report_format
    worker_limit = scan_report_obj.worker_limit
    filter_inputs = scan_report_obj.filter_inputs

    extension = "xlsx" if report_format.lower() == "excel" else "csv"
    report_name = f"scan_report_{report_id}.{extension}"
    report_path = os.path.join(settings.MEDIA_ROOT, report_name)

    # Getting connector object.
    connector_obj = get_connector_obj(
        connector_type=CONNECTOR_TYPE,
        private_ip=PRIVATE_IP,
        customer_db=customer_db,
    )

    scan_report_obj_up.update(
        connector_obj=connector_obj,
        report_name=report_name, 
        report_status="Processing"
    )

    filtered_result_objs = filter_result_objs(customer_db, filter_inputs)
    total_results = filtered_result_objs.count()
    chunk_size = settings.REPORT_CHUNK_SIZE
    progress_per_chunk = 100 / (total_results // chunk_size + 1) if total_results > chunk_size else 100

    scan_report_obj_up.update(report_status="Writing")

    assigned_tasks = []
    report_chunks = []
    progress = 0
    for chunked_ids in get_chunked_ids(filtered_result_objs, chunk_size=chunk_size):
        task = write_scan_report_chunk.delay(
            customer_db=customer_db,
            report_id=report_id,
            chunked_ids=chunked_ids,
            tmp_dir=settings.MEDIA_ROOT,
        )
        assigned_tasks.append(task)

        while len(assigned_tasks) >= worker_limit:
            for task in assigned_tasks.copy():
                try:
                    if task.ready():
                        report_chunks.append(task.get())
                        assigned_tasks.remove(task)
                        progress += progress_per_chunk
                        print(f"Progress: {progress}%")
                except Exception as e:
                    print(f"Error processing task {task.id}: {e}")
                    assigned_tasks.remove(task)
                    progress += progress_per_chunk
            
            report_progress = round(progress, 2)
            scan_report_obj_up.update(report_progress=report_progress)
            time.sleep(1)

    # Handle any remaining tasks
    while len(assigned_tasks) > 0:
        for task in assigned_tasks.copy():
            try:
                if task.ready():
                    report_chunks.append(task.get())
                    assigned_tasks.remove(task)
                    progress += progress_per_chunk
                    print(f"Progress: {progress}%")
            except Exception as e:
                print(f"Error processing task {task.id}: {e}")
                assigned_tasks.remove(task)
                progress += progress_per_chunk
        report_progress = round(progress, 2)
        scan_report_obj_up.update(report_progress=report_progress)
        time.sleep(1)

    
    scan_report_obj_up.update(report_status="Merging")

    # Merging all chunked reports into a single file.
    merge_report_chunks(
        customer_db=customer_db,
        report_id=report_id,
        report_chunks=report_chunks,
        merged_report_path=report_path,
        sheet_title=f"Scan Report"
    )

    # Final updating the report object.
    scan_report_obj_up.update(
        report_path=report_path,
        report_status="Generated",
    )

    print(f"Scan report generated: {report_path}")
    return


@shared_task
def write_data_report_chunk(customer_db, report_id, chunked_ids, tmp_dir):
    """
    Builds a pickle with DataFrame columns:
      ["Data", "Data Type", "Department", "Platforms"]

    - Accepts EITHER Result IDs OR Data IDs in `chunked_ids`
      (auto-detects which one)
    - One row per unique data_obj (data_obj__id)
    - Masked/Unmasked decided by GenReport.mask_status:
        masked   -> data_obj__data
        unmasked -> decrypt(data_obj__value_obj__encrypted_value)
    - Department: "Dept:count | Dept:count"
    - Platforms : "Platform:count, Platform:count"
    """
    try:
        update_celery_config(customer_dbs_to_update=[customer_db])
    except Exception:
        pass

    first_id = chunked_ids[0] if chunked_ids else "empty"
    last_id  = chunked_ids[-1] if chunked_ids else "empty"
    out_path = os.path.join(tmp_dir, f"data_report_chunk_{customer_db}_{report_id}_{first_id}-{last_id}.pkl")
    os.makedirs(tmp_dir, exist_ok=True)
    print(f"[write_data_report_chunk] writing: {out_path} (ids_count={len(chunked_ids)})")

    cols = ["Data", "Data Type", "Department", "Platforms"]

    if not chunked_ids:
        pd.DataFrame(columns=cols).to_pickle(out_path)
        print(f"[write_data_report_chunk] empty chunk written: {out_path}")
        return out_path

    # Mask choice decides which field to display (do NOT filter on is_masked)
    mask_status = (
        GenReport.objects.using(customer_db).get(id=report_id).mask_status
    )
    want_masked = (str(mask_status).strip().lower() == "masked")
    print(f"[write_data_report_chunk] mask_status={mask_status} -> want_masked={want_masked}")

    # --- AUTO-DETECT ID TYPE ---
    # Try treating chunked_ids as Result PKs
    result_qs = Result.objects.using(customer_db).filter(pk__in=chunked_ids)

    # Pull everything needed from Result only
    result_rows = list(
        result_qs.annotate(
            rid=F("id"),
            data_id=F("data_obj__id"),
            masked_data=F("data_obj__data"),
            enc_val=F("data_obj__value_obj__encrypted_value"),
            dtype=F("data_obj__data_type_obj__data_type"),
            dept=F("host_obj__department_objs__department_name"),
            plat=F("host_obj__platform_type"),
        )
        .values("rid", "data_id", "masked_data", "enc_val", "dtype", "dept", "plat")
        .distinct()
    )

    if not result_rows:
        pd.DataFrame(columns=cols).to_pickle(out_path)
        print(f"[write_data_report_chunk] no Result rows after auto-detect → {out_path}")
        return out_path

    # -------- Build dictionary keyed by data_id --------
    data_set = {}  # data_id -> {"Data","Data Type","Departments":{..},"Platforms":{..}}
    seen_department = set() 
    seen_platform = set()  

    for row in result_rows:
        data_id = row["data_id"]
        if data_id is None:
            continue

        # resolve Data once per data_id according to mask
        if data_id not in data_set:
            if want_masked:
                v = row.get("masked_data")
                data_val = v.strip() if isinstance(v, str) else (str(v).strip() if v is not None else "N/A")
            else:
                ev = row.get("enc_val")
                if ev:
                    try:
                        dv = decrypt_data(ev)
                    except Exception:
                        dv = ""
                else:
                    dv = ""
                data_val = dv.strip() if isinstance(dv, str) else (str(dv).strip() if dv is not None else "N/A")

            dt = row.get("dtype")
            dtype_val = dt.strip() if isinstance(dt, str) else (str(dt).strip() if dt is not None else "N/A")

            data_set[data_id] = {
                "Data": data_val if data_val else "N/A",
                "Data Type": dtype_val if dtype_val else "N/A",
                "Departments": defaultdict(int),
                "Platforms": defaultdict(int),
            }
        # aggregate Departments (dedupe per result row)
        dept = row.get("dept")
        dept_name = dept.strip() if isinstance(dept, str) else (str(dept).strip() if dept is not None else "")
        if dept_name:
            tok = (data_id, dept_name, row["rid"])
            if tok not in seen_department:
                data_set[data_id]["Departments"][dept_name] += 1
                seen_department.add(tok)

        # aggregate Platforms (dedupe per result row)
        plat = row.get("plat")
        plat_name = plat.strip() if isinstance(plat, str) else (str(plat).strip() if plat is not None else "")
        if plat_name:
            tok = (data_id, plat_name, row["rid"])
            if tok not in seen_platform:
                data_set[data_id]["Platforms"][plat_name] += 1
                seen_platform.add(tok)

    # -------- Convert dict -> rows -> DataFrame --------
    rows = []
    for data_id in sorted(data_set.keys(), reverse=True):
        data = data_set[data_id]
        dept_map = data["Departments"]
        plat_map = data["Platforms"]

        dept_str = " | ".join(
            f"{k}:{dept_map[k]}" for k in sorted(dept_map, key=lambda n: (-dept_map[n], n))
        ) if dept_map else "N/A"

        platforms_str = ", ".join(
            f"{k}:{plat_map[k]}" for k in sorted(plat_map, key=lambda n: (-plat_map[n], n))
        ) if plat_map else ""  # you said platform should always exist; keep blank if anomaly

        rows.append({
            "Data": data["Data"],
            "Data Type": data["Data Type"],
            "Department": dept_str,
            "Platforms": platforms_str,
        })

    df = pd.DataFrame(rows, columns=cols)
    for c in ["Data", "Data Type", "Department"]:
        df[c] = df[c].replace("", "N/A").fillna("N/A")

    df.to_pickle(out_path)
    print(f"[write_data_report_chunk] final DataFrame shape={df.shape}  rows={len(df)}  unique_data_ids={len(data_set)}")
    return out_path

# unused
def generate_data_report_old(customer_db: str, report_id: int):
    """
    Generates report of scan results.
    """

    data_report_obj_up = GenReport.objects.using(customer_db).filter(id=int(report_id))
    data_report_obj = data_report_obj_up.first()
    print(f'data_report_obj ==> {data_report_obj.__dict__}')
    report_id = data_report_obj.id
    report_format = data_report_obj.report_format
    worker_limit = data_report_obj.worker_limit
    filter_inputs = data_report_obj.filter_inputs

    extension = "xlsx" if report_format.lower() == "excel" else "csv"
    report_name = f"data_report_{report_id}.{extension}"
    report_path = os.path.join(settings.MEDIA_ROOT, report_name)

    # Getting connector object.
    connector_obj = get_connector_obj(
        connector_type=CONNECTOR_TYPE,
        private_ip=PRIVATE_IP,
        customer_db=customer_db,
    )

    data_report_obj_up.update(
        connector_obj=connector_obj,
        report_name=report_name, 
        report_status="Processing"
    )

    filtered_result_objs = filter_result_objs(customer_db, filter_inputs)
    data_ids = filtered_result_objs.values_list('data_obj__id', flat=True).distinct()
    filtered_data_objs = Data.objects.using(customer_db).filter(id__in=data_ids)
    total_data = filtered_data_objs.count()
    chunk_size = settings.REPORT_CHUNK_SIZE
    progress_per_chunk = 100 / (total_data // chunk_size + 1) if total_data > chunk_size else 100

    data_report_obj_up.update(report_status="Writing")

    assigned_tasks = []
    report_chunks = []
    progress = 0
    for chunked_ids in get_chunked_ids(filtered_data_objs, chunk_size=chunk_size):
        task = write_data_report_chunk.delay(
            customer_db=customer_db,
            report_id=report_id,
            chunked_ids=chunked_ids,
            tmp_dir=settings.MEDIA_ROOT,
        )
        assigned_tasks.append(task)
        print(f"assigned_tasks==============={assigned_tasks}")

        while len(assigned_tasks) >= worker_limit:
            for task in assigned_tasks.copy():
                try:
                    if task.ready():
                        report_chunks.append(task.get())
                        assigned_tasks.remove(task)
                        progress += progress_per_chunk
                        print(f"Progress: {progress}%")
                except Exception as e:
                    print(f"Error processing task {task.id}: {e}")
                    assigned_tasks.remove(task)
                    progress += progress_per_chunk
            
            report_progress = round(progress, 2)
            data_report_obj_up.update(report_progress=report_progress)
            time.sleep(1)

    # Handle any remaining tasks
    while len(assigned_tasks) > 0:
        for task in assigned_tasks.copy():
            try:
                if task.ready():
                    report_chunks.append(task.get())
                    assigned_tasks.remove(task)
                    progress += progress_per_chunk
                    print(f"Progress: {progress}%")
            except Exception as e:
                print(f"Error processing task {task.id}: {e}")
                assigned_tasks.remove(task)
                progress += progress_per_chunk
        report_progress = round(progress, 2)
        data_report_obj_up.update(report_progress=report_progress)
        time.sleep(1)

    
    data_report_obj_up.update(report_status="Merging")

    # Merging all chunked reports into a single file.
    merge_report_chunks(
        customer_db=customer_db,
        report_id=report_id,
        report_chunks=report_chunks,
        merged_report_path=report_path,
        sheet_title=f"Data Report"
    )

    # Final updating the report object.
    data_report_obj_up.update(
        report_path=report_path,
        report_status="Generated",
    )

    print(f"Unique data report generated: {report_path}")
    return


@shared_task
def write_table_report_chunk_old(customer_db, report_id, chunked_ids, tmp_dir):
    """
    Write a chunk of ColumnName data to a temporary pickle file.
    """
    update_celery_config(customer_dbs_to_update=[customer_db])
    
    # Building the temporary report chunk path.
    report_chunk = os.path.join(
        tmp_dir,
        f"data_report_chunk_{customer_db}_{report_id}_{chunked_ids[0]}-{chunked_ids[-1]}.pkl"
    )
    print(f'report_chunk: {report_chunk}')

    report_obj = GenReport.objects.using(customer_db).get(id=report_id)
    mask_status = report_obj.mask_status

    if mask_status == "masked":
        field_names = [
            "data",
            "data_type_obj__data_type",
            "data_type_obj__subclassification_objs__subclassification",
            "data_type_obj__subclassification_objs__classification_obj__classification"
        ]
    elif mask_status == "unmasked":
        field_names = [
            "value_obj__encrypted_value",
            "data_type_obj__data_type",
            "data_type_obj__subclassification_objs__subclassification",
            "data_type_obj__subclassification_objs__classification_obj__classification"
        ]
    else:
        raise ValueError(f"Invalid mask status: {mask_status}. Expected 'masked' or 'unmasked'.")

    data_objs_chunk = Data.objects.using(customer_db).filter(pk__in=chunked_ids)
    
    # Fetching rows in a list.
    data_objs_list = list(data_objs_chunk.values(*field_names))

    if mask_status == "unmasked":
        for data_obj in data_objs_list:
            encrypted_data = data_obj["value_obj__encrypted_value"]
            decrypted_data = decrypt_data(encrypted_data)
            data_obj["data"] = decrypted_data
            del data_obj["value_obj__encrypted_value"]


    print(f'Creating DataFrame with {len(data_objs_list)} rows')
    df = pd.DataFrame(data_objs_list)
    print(f'DataFrame created with shape: {df.shape}')

    # 4) Rename columns to your desired field names
    df.rename(
        columns={
            "data": "Data",
            "data_type_obj__data_type": "Data Type",
            "data_type_obj__subclassification_objs__subclassification": "Subclassifications",
            "data_type_obj__subclassification_objs__classification_obj__classification": "Classifications",
        },
        inplace=True
    )
    print(f'Renaming columns completed. DataFrame shape: {df.shape}')

    desired_order = [
        "Data", "Data Type",
        "Subclassifications", "Classifications"
    ]

    # This returns a new DataFrame with exactly that column sequence:
    df = df[desired_order]

    # serialize to pickle
    df.to_pickle(report_chunk)

    print(f'Chunk pickle written to: {report_chunk}')
    return report_chunk



@shared_task
def write_table_report_chunk(customer_db, report_id, chunked_ids, tmp_dir):
    """
    Write a chunk of TableName data to a temporary pickle file.
    """
    update_celery_config(customer_dbs_to_update=[customer_db])
    # Building the temporary chunked report path.
    report_chunk = os.path.join(
        tmp_dir,
        f"chunked_table_report_{customer_db}_{report_id}_{chunked_ids[0]}-{chunked_ids[-1]}.pkl"
    )
    print(f'report_chunk: {report_chunk}')

    chunked_table_name_objs = TableName.objects.using(customer_db).filter(pk__in=chunked_ids)
    # Fetching rows in a list.
    table_names_list = list(
        chunked_table_name_objs
        .values(
            "id",
            "table_name",
            "db_name_obj__db_name",
            "db_name_obj__host_obj__id",
            "db_name_obj__host_obj__scan_host",
            "db_name_obj__host_obj__scan_platform",
            "scan_status",
            "cell_count",
        )
    )
    sb_totals = (
        ScanBreakdown.objects.using(customer_db)
        .filter(scan_unit_obj__table_name_obj__in=chunked_ids, data_count__gt=0)
        .values('scan_unit_obj__table_name_obj__id')
        .annotate(total=Sum('data_count'))
    )
    totals_by_table = {d['scan_unit_obj__table_name_obj__id']: d['total'] for d in sb_totals}
    host_ids = {d['db_name_obj__host_obj__id'] for d in table_names_list}
    hosts = Host.objects.using(customer_db).filter(id__in=host_ids).prefetch_related('host_tags')
    host_map = {h.id: h for h in hosts}
    for table_name_dict in table_names_list:
        table_id = table_name_dict["id"]
        table_name_dict['data_found'] = totals_by_table.get(table_id, 0)
        host_id = table_name_dict["db_name_obj__host_obj__id"]
        host_obj = host_map[host_id]
        host_tags_str = ', '.join(t.tag_name for t in host_obj.host_tags.all()) or 'N/A'
        table_name_dict["host_tags"] = host_tags_str
        if not table_name_dict["scan_status"]:
            table_name_dict["scan_status"] = "Not Scanned Yet"
        # Removing some keys which is not required in final DataFrame.
        del table_name_dict["id"]
        del table_name_dict["db_name_obj__host_obj__id"]
    print(f'Creating DataFrame with {len(table_names_list)} rows')
    df = pd.DataFrame(table_names_list)
    print(f'DataFrame created with shape: {df.shape}')
    # 4) Rename columns to your desired field names
    df.rename(
        columns={
            "table_name": "Table",
            "db_name_obj__db_name": "Database",
            "db_name_obj__host_obj__scan_host": "Host",
            "db_name_obj__host_obj__scan_platform": "Platform",
            "host_tags": "Host Tags",
            "scan_status": "Scan Status",
            "data_found": "Data Found",
            "cell_count": "Cell Count",
        },
        inplace=True
    )
    print(f'Renaming columns completed. DataFrame shape: {df.shape}')
    desired_order = [
        "Table", "Database",
        "Host", "Platform", "Host Tags",
        "Scan Status", "Data Found", "Cell Count"
    ]
    # This returns a new DataFrame with exactly that column sequence:
    df = df[desired_order]
    # serialize to pickle
    df.to_pickle(report_chunk)
    print(f'Chunked pickle written to: {report_chunk}')
    return report_chunk


@shared_task
def write_scan_unit_report_chunk(customer_db, report_id, chunked_ids, tmp_dir):
    """
    Write a chunk of ScanUnit data to a temporary pickle file.
    Works like write_table_report_chunk but for ScanUnit/GenReport.
    """
    update_celery_config(customer_dbs_to_update=[customer_db])
    print("Dispatching chunk size:", len(chunked_ids))

    # Chunked file path
    report_chunk = os.path.join(
        tmp_dir,
        f"chunked_scan_unit_report_{customer_db}_{report_id}_{chunked_ids[0]}-{chunked_ids[-1]}.pkl"
    )
    print(f'report_chunk: {report_chunk}')

    # Fetch ScanUnit rows
    scan_units_qs = ScanUnit.objects.using(customer_db).filter(pk__in=chunked_ids)

    scan_units_list = list(
        scan_units_qs.values(
            "id",
            "host_obj__id",
            "host_obj__scan_host",
            "host_obj__scan_platform",
            "host_obj__platform_type",
            "file_path",
            "created_on",
            "created_by",
            "modified_on",
            "modified_by",
            "table_name",
            "db_name",
            "column_name",
            "email_label",
            "email_subject",
            "email_sender",
            "email_receivers",
            "tagging_status",
            "patching_statuses__patching_status"
        )
    )

    # Prefetch host_tags + departments
    host_ids = {su["host_obj__id"] for su in scan_units_list}
    hosts = Host.objects.using(customer_db).filter(id__in=host_ids).prefetch_related("host_tags", "department_objs")
    host_map = {h.id: h for h in hosts}

    # Format scan_unit data
    for su_dict in scan_units_list:
        host_id = su_dict["host_obj__id"]
        host_obj = host_map[host_id]
        su_dict["host_tags"] = ", ".join(t.tag_name for t in host_obj.host_tags.all()) or "N/A"
        su_dict["department"] = ", ".join(d.department_name for d in host_obj.department_objs.all()) or "N/A"

        if su_dict["host_obj__platform_type"] in DRIVE_PLATFORMS or su_dict["host_obj__platform_type"] in CLOUD_PLATFORMS:
            # File info
            file_info = f"File Path: {su_dict['file_path'] or 'N/A'}"
            if su_dict["created_on"]:
                file_info += f" | Created: {su_dict['created_on']}"
            if su_dict["created_by"]:
                file_info += f" | Created By: {su_dict['created_by']}"
            if su_dict["modified_on"]:
                file_info += f" | Modified: {su_dict['modified_on']}"
            if su_dict["modified_by"]:
                file_info += f" | Modified By: {su_dict['modified_by']}"
            su_dict["data_location"] = file_info
            su_dict["patching_status"] = su_dict['patching_statuses__patching_status'] or "Not Patched"
        elif su_dict["host_obj__platform_type"] in DB_PLATFORMS:
            # Database platforms
            db_info = f"DB Name: {su_dict['db_name']}, Table: {su_dict['table_name']}, Column: {su_dict['column_name']}"
            su_dict["data_location"] = db_info
            su_dict["patching_status"] = "N/A"
        elif su_dict["host_obj__scan_platform"] in EMAIL_PLATFORMS:
            # Email platforms
            email_info = f"Label: {su_dict['email_label'] or 'N/A'}, Subject: {su_dict['email_subject'] or 'N/A'}, "\
                         f"Sender: {su_dict['email_sender'] or 'N/A'}, Receivers: {su_dict['email_receivers'] or 'N/A'}"
            su_dict["data_location"] = email_info
            su_dict["patching_status"] = "N/A"

        # Cleanup unneeded keys
        for key in ["id", "host_obj__id", "file_path", "table_name", "db_name", "column_name",
                    "created_on", "created_by", "modified_on", "modified_by", "host_obj__platform_type"]:
            su_dict.pop(key, None)

    print(f'Creating DataFrame with {len(scan_units_list)} rows')

    df = pd.DataFrame(scan_units_list)

    # Rename columns
    df.rename(
        columns={
            "host_obj__scan_host": "Host",
            "host_obj__scan_platform": "Platform",
            "host_tags": "Host Tags",
            "department": "Department",
            "data_location": "Data Location",
            "patching_status": "Patching Status",
            "tagging_status": "Tagging Status",
        },
        inplace=True
    )

    desired_order = [
        "Host", "Platform", "Host Tags", "Department",
        "Data Location", "Patching Status", "Tagging Status"
    ]
    df = df[desired_order]
    print(f'final DataFrame shape: {df.shape}')

    # Write pickle
    df.to_pickle(report_chunk)
    print(f'Chunked pickle written to: {report_chunk}')
    return report_chunk


@shared_task
def write_found_data_type_report_chunk(customer_db, report_id, chunked_ids, tmp_dir):
    """
    Build a single-row-per Data Type summary:
      - Classification, Sub-Classification, Department -> merged unique values or "N/A"
      - Platform Wise Data -> "Platform:total" items separated by " | "
    Writes a pickle file and returns its path.
    """
    update_celery_config(customer_dbs_to_update=[customer_db])
    print("Dispatching chunk size:", len(chunked_ids))

    report_chunk = os.path.join(
        tmp_dir,
        f"chunked_found_data_type_report_{customer_db}_{report_id}_{chunked_ids[0]}-{chunked_ids[-1]}.pkl"
    )
    print(f"report_chunk: {report_chunk}")

    sb_rows = list(
        ScanBreakdown.objects.using(customer_db)
        .filter(data_type_obj__in=chunked_ids, data_count__gt=0)
        .annotate(
            data_type=F("data_type_obj__data_type"),
            subclassification=F("data_type_obj__subclassification_objs__subclassification"),
            classification=F("data_type_obj__subclassification_objs__classification_obj__classification"),
            platform=F("host_obj__platform_type"),
            department=F("host_obj__department_objs__department_name"),
        )
        .values("data_type", "classification", "subclassification", "platform", "department")
        .annotate(row_count=Count("id"))   # lightweight tally used for platform totals
    )

    # Empty guard → write empty frame with expected columns
    columns = ["Data Type", "Classification", "Sub-Classification", "Department", "Platform Wise Data"]
    if not sb_rows:
        pd.DataFrame(columns=columns).to_pickle(report_chunk)
        print(f"Chunked pickle written to: {report_chunk} (empty)")
        return report_chunk

    df = pd.DataFrame(sb_rows).rename(columns={
        "data_type": "Data Type",
        "classification": "Classification",
        "subclassification": "Sub-Classification",
        "platform": "Platform",
        "department": "Department",
        "row_count": "Count",
    })

    # Merge unique values helper (inline)
    merge_pipe = lambda s: " | ".join(sorted({str(x).strip() for x in s if pd.notna(x) and str(x).strip()})) or "N/A"

    # A) One-row-per Data Type: merged strings for the three descriptive cols
    base_df = (
        df.groupby("Data Type", as_index=False)
          .agg({
              "Classification": merge_pipe,
              "Sub-Classification": merge_pipe,
              "Department": merge_pipe,
          })
    )

    # B) Platform Wise Data: sum counts per (Data Type, Platform) → format as "Platform:total"
    platform_wise_data_type_count = (
        df.groupby(["Data Type", "Platform"], dropna=False)["Count"]
          .sum()
          .reset_index()
    )
    # format per dtype: sort by total desc then platform asc, join
    platform_str = (
        platform_wise_data_type_count.sort_values(["Data Type", "Count", "Platform"], ascending=[True, False, True])
                  .groupby("Data Type", as_index=False)
                  .apply(lambda g: " | ".join(
                      f"{p}:{int(c)}" for p, c in g[["Platform", "Count"]].itertuples(index=False, name=None)
                      if pd.notna(p) and str(p).strip()
                  ) or "N/A")
                  .rename(columns={None: "Platform Wise Data"})
    )

    # C) Merge and order columns
    final_df = base_df.merge(platform_str, on="Data Type")
    final_df = final_df[columns]

    # Write
    os.makedirs(tmp_dir, exist_ok=True)
    final_df.to_pickle(report_chunk)
    print(f"Chunked pickle written to: {report_chunk}")
    return report_chunk


@shared_task
def write_host_user_detail_report_chunk(customer_db, report_id, chunked_ids, tmp_dir):
    """
    Write a chunk of HostUser data to a temporary pickle file.
    """
    update_celery_config(customer_dbs_to_update=[customer_db])

    # Chunked file path
    report_chunk = os.path.join(
        tmp_dir,
        f"chunked_host_user_detail_report_{customer_db}_{report_id}_{chunked_ids[0]}-{chunked_ids[-1]}.pkl"
    )
    print(f'report_chunk: {report_chunk}')
    # Fetch HostUser rows
    chunked_host_user_detail_objs = HostUser.objects.using(customer_db).filter(pk__in=chunked_ids)

    host_user_obj_list = list(
        chunked_host_user_detail_objs.values(
            "id",
            "username",
            "host_obj__scan_host",
            "host_obj__platform_type",
            "host_obj__scan_platform",
            "host_obj__host_tags__tag_name",
            "host_obj__department_objs__department_name",
        )
    )

    for host_user_obj in host_user_obj_list:
        host_tags = host_user_obj.get("host_obj__host_tags__tag_name")
        host_user_obj["host_tags"] = ", ".join(sorted(set(host_tags.split(", ")))) if host_tags else "N/A"

        department = host_user_obj.get("host_obj__department_objs__department_name")
        host_user_obj["department"] = ", ".join(sorted(set(department.split(", ")))) if department else "N/A"

        permissions = HostUserPermission.objects.using(customer_db).filter(host_user_obj__username=host_user_obj['username']).count()
        host_user_obj["permissions"] = permissions


    
    # Build DataFrame
    df = pd.DataFrame(host_user_obj_list)

    # Correct, consistent rename -> final report headers
    df.rename(
        columns={
            "username": "Username",
            "host_obj__scan_host": "Host",
            "host_obj__scan_platform": "Platform",
            "host_obj__platform_type": "Platform Type",  # keep if you want it too
            "host_tags": "Host Tags",
            "department": "Department",
            "permissions": "Permission",
        },
        inplace=True,
    )

    # Choose the final column order you want in the report
    desired_order = [
        "Username", "Host", "Platform", "Host Tags", "Department", "Permission"
    ]

    # If you also want 'Platform Type' in the report, add it to desired_order.

    # Reindex to avoid KeyError if any column is missing (will fill with NaN)
    # df = df.reindex(columns=desired_order)
    # Ensure all desired columns exist (create empty ones if missing)
    for col in desired_order:
        if col not in df.columns:
            df[col] = pd.NA
    df = df[desired_order]

    # Write pickle
    df.to_pickle(report_chunk)
    print(f'Chunked pickle written to: {report_chunk}')
    return report_chunk


@shared_task
def write_scan_details_report_chunk(customer_db, report_id, chunked_ids, tmp_dir):
    """
    Write a chunk of ScanDetails data to a temporary pickle file.
    """
    update_celery_config(customer_dbs_to_update=[customer_db])

    # Chunked file path
    report_chunk = os.path.join(
        tmp_dir,
        f"chunked_scan_details_report_{customer_db}_{report_id}_{chunked_ids[0]}-{chunked_ids[-1]}.pkl"
    )
    print(f'report_chunk: {report_chunk}')

    # Fetch ScanDetails rows
    chunked_scan_details_objs = ScanDetails.objects.using(customer_db).filter(pk__in=chunked_ids)

    scan_details_list = list(
        chunked_scan_details_objs.values(
            "id",
            "scan_obj__id",
            "scan_unit_obj__host_obj__id",
            "scan_unit_obj__host_obj__scan_host",
            "scan_unit_obj__host_obj__scan_platform",
            "scan_unit_obj__host_obj__platform_type",
            "scan_start_time",
            "scan_end_time",
            "scan_pause_time",
            "scan_status",
            "scan_unit_obj__file_path",
            "scan_unit_obj__table_name",
            "scan_unit_obj__db_name",
            "scan_unit_obj__column_name",
            "scan_unit_obj__email_subject",
            "scan_unit_obj__email_label",
            "scan_unit_obj__email_sender",
            "scan_unit_obj__email_receivers",
            "scan_unit_obj__content_size",
            "tagging_status",
            "data_found"
        )
    )

    # Get data count totals from ScanBreakdown
    sb_totals = (
        ScanBreakdown.objects.using(customer_db)
        .filter(scan_details_obj__in=chunked_ids, data_count__gt=0)
        .values('scan_details_obj__id')
        .annotate(total=Sum('data_count'))
    )
    totals_by_scan_details = {d['scan_details_obj__id']: d['total'] for d in sb_totals}

    # Prefetch host + tags + departments
    host_ids = {p["scan_unit_obj__host_obj__id"] for p in scan_details_list if p.get("scan_unit_obj__host_obj__id")}
    hosts = Host.objects.using(customer_db).filter(id__in=host_ids).prefetch_related("host_tags", "department_objs")
    host_map = {h.id: h for h in hosts}

    for details_dict in scan_details_list:
        scan_details_id = details_dict["id"]
        details_dict['data_count'] = totals_by_scan_details.get(scan_details_id, 0)
        
        host_id = details_dict.get("scan_unit_obj__host_obj__id")
        if host_id and host_id in host_map:
            host_obj = host_map[host_id]
            details_dict["host_tags"] = ", ".join(t.tag_name for t in host_obj.host_tags.all()) or "N/A"
            details_dict["department"] = ", ".join(d.department_name for d in host_obj.department_objs.all()) or "N/A"
        else:
            details_dict["host_tags"] = "N/A"
            details_dict["department"] = "N/A"

        # Data Location logic
        if details_dict["scan_unit_obj__host_obj__platform_type"] in DRIVE_PLATFORMS:
            # Drive
            details_dict["data_location"] = f"File Path: {details_dict['scan_unit_obj__file_path'] or 'N/A'}"
        elif details_dict["scan_unit_obj__db_name"]:
            # Database
            db_info = f"DB Name: {details_dict['scan_unit_obj__db_name']} | Table: {details_dict['scan_unit_obj__table_name'] or 'N/A'} | Column Name: {details_dict['scan_unit_obj__column_name'] or 'N/A'}"
            details_dict["data_location"] = db_info
        elif details_dict["scan_unit_obj__host_obj__platform_type"] in EMAIL_PLATFORMS:
            details_dict["data_location"] = f"Email Subject: {details_dict['scan_unit_obj__email_subject']} | Email Label: {details_dict['scan_unit_obj__email_label']} | Email Sender: {details_dict['scan_unit_obj__email_sender']} | Email Recievers: {details_dict['scan_unit_obj__email_receivers']}"
            
        else:
            details_dict["data_location"] = "N/A"

        # Set scan status if empty
        if not details_dict["scan_status"]:
            details_dict["scan_status"] = "Not Scanned Yet"

        # Create combined Scan Details column
        scan_details_info = []
        
        # Add scan status
        if details_dict["scan_obj__id"]:
            scan_details_info.append(f"Scan ID: {details_dict['scan_obj__id']}")
        if details_dict["scan_status"]:
            scan_details_info.append(f"Status: {details_dict['scan_status']}")
        
        # Add scan start time
        if details_dict["scan_start_time"]:
            scan_details_info.append(f"Started: {details_dict['scan_start_time']}")
        
        # Add scan end time
        if details_dict["scan_end_time"]:
            scan_details_info.append(f"Ended: {details_dict['scan_end_time']}")
        
        # Add scan pause time if available
        if details_dict["scan_pause_time"]:
            scan_details_info.append(f"Paused: {details_dict['scan_pause_time']}")
        
        
        # Combine all scan details
        details_dict["scan_details"] = " | ".join(scan_details_info) if scan_details_info else "N/A"

        # Cleanup unneeded keys
        for key in [
            "id", "scan_obj__id", "scan_unit_obj__host_obj__id", "scan_unit_obj__host_obj__platform_type",
            "scan_unit_obj__file_path", "scan_unit_obj__db_name", 
            "scan_unit_obj__table_name", "scan_unit_obj__column_name",
            "scan_status", "scan_start_time", 
            "scan_end_time", "scan_pause_time", "data_found"
        ]:
            details_dict.pop(key, None)

    print(f'Creating DataFrame with {len(scan_details_list)} rows')

    df = pd.DataFrame(scan_details_list)

    # Rename columns
    df.rename(
        columns={
            "scan_unit_obj__host_obj__scan_host": "Host",
            "scan_unit_obj__host_obj__scan_platform": "Platform",
            "host_tags": "Host Tags",
            "department": "Department",
            "scan_details": "Scan Details",
            "data_location": "Data Location",
            "data_count": "Data Count",
            "tagging_status": "Tagging Status",
        },
        inplace=True
    )

    desired_order = [
        "Host", "Platform", "Host Tags", "Department", "Scan Details", 
        "Data Location", "Tagging Status", "Data Count"
    ]
    df = df[desired_order]

    # Write pickle
    df.to_pickle(report_chunk)
    print(f'Chunked pickle written to: {report_chunk}')
    return report_chunk


@shared_task
def write_db_report_chunk(customer_db, report_id, chunked_ids, tmp_dir):
    """
    Write a chunk of DbName data to a temporary pickle file.
    """
    update_celery_config(customer_dbs_to_update=[customer_db])
    # Building the temporary chunked report path.
    report_chunk = os.path.join(
        tmp_dir,
        f"chunked_db_report_{customer_db}_{report_id}_{chunked_ids[0]}-{chunked_ids[-1]}.pkl"
    )
    print(f'report_chunk: {report_chunk}')
    chunked_db_name_objs = DBName.objects.using(customer_db).filter(pk__in=chunked_ids)
    # Fetching rows in a list.
    db_names_list = list(
        chunked_db_name_objs
        .values(
            "id",
            "db_name",
            "host_obj__id",
            "host_obj__scan_host",
            "host_obj__scan_platform",
            "scan_status",
            "cell_count",
        )
    )
    sb_totals = (
        ScanBreakdown.objects.using(customer_db)
        .filter(scan_unit_obj__db_name_obj__in=chunked_ids, data_count__gt=0)
        .values('scan_unit_obj__db_name_obj__id')
        .annotate(total=Sum('data_count'))
    )
    totals_by_db = {d['scan_unit_obj__db_name_obj__id']: d['total'] for d in sb_totals}
    host_ids = {d['host_obj__id'] for d in db_names_list}
    hosts = Host.objects.using(customer_db).filter(id__in=host_ids).prefetch_related('host_tags')
    host_map = {h.id: h for h in hosts}
    for db_name_dict in db_names_list:
        db_id = db_name_dict["id"]
        db_name_dict['data_found'] = totals_by_db.get(db_id, 0)
        host_id = db_name_dict["host_obj__id"]
        host_obj = host_map[host_id]
        host_tags_str = ', '.join(t.tag_name for t in host_obj.host_tags.all()) or 'N/A'
        db_name_dict["host_tags"] = host_tags_str
        if not db_name_dict["scan_status"]:
            db_name_dict["scan_status"] = "Not Scanned Yet"
        # Removing some keys which is not required in final DataFrame.
        del db_name_dict["id"]
        del db_name_dict["host_obj__id"]
    print(f'Creating DataFrame with {len(db_names_list)} rows')
    df = pd.DataFrame(db_names_list)
    print(f'DataFrame created with shape: {df.shape}')
    # 4) Rename columns to your desired field names
    df.rename(
        columns={
            "db_name": "Database",
            "host_obj__scan_host": "Host",
            "host_obj__scan_platform": "Platform",
            "host_tags": "Host Tags",
            "scan_status": "Scan Status",
            "data_found": "Data Found",
            "cell_count": "Cell Count",
        },
        inplace=True
    )
    print(f'Renaming columns completed. DataFrame shape: {df.shape}')
    desired_order = [
        "Database", "Host",
        "Platform", "Host Tags",
        "Scan Status", "Data Found", "Cell Count"
    ]
    # This returns a new DataFrame with exactly that column sequence:
    df = df[desired_order]
    # serialize to pickle
    df.to_pickle(report_chunk)
    print(f'Chunked pickle written to: {report_chunk}')
    return report_chunk


@shared_task
def write_bucket_name_report_chunk(customer_db, report_id, chunked_ids, tmp_dir):
    """
    Write a chunk of BucketName data to a temporary pickle file.
    """
    update_celery_config(customer_dbs_to_update=[customer_db])
    # Building the temporary chunked report path.
    report_chunk = os.path.join(
        tmp_dir,
        f"chunked_bucket_report_{customer_db}_{report_id}_{chunked_ids[0]}-{chunked_ids[-1]}.pkl"
    )
    print(f'report_chunk: {report_chunk}')

    chunked_bucket_name_objs = BucketName.objects.using(customer_db).filter(pk__in=chunked_ids)

    # Fetching rows in a list.
    bucket_names_list = list(  
        chunked_bucket_name_objs
        .values(
            "id",
            "bucket_name",
            "region",
            "host_obj__id",
            "host_obj__scan_host",
            "is_public",
            "is_encrypted",
            "object_count",
            "scan_status",

        )
    )

    host_ids = {d['host_obj__id'] for d in bucket_names_list if d.get('host_obj__id')}

    hosts = Host.objects.using(customer_db).filter(id__in=host_ids).prefetch_related('host_tags', 'department_objs')
    host_map = {h.id: ', '.join(t.tag_name for t in h.host_tags.all()) if h.host_tags else 'N/A' for h in hosts}
    department_map = {h.id: ', '.join(d.department_name for d in h.department_objs.all()) or 'N/A' for h in hosts}

    for bucket_name_dict in bucket_names_list:
        bucket_id = bucket_name_dict["id"]
        host_id = bucket_name_dict.get("host_obj__id")
        host_obj = host_map.get(host_id)
        # host_tags_str = ', '.join(t.tag_name for t in host_obj.host_tags.all()) if host_obj else 'N/A'
        bucket_name_dict["host_tags"] = host_map.get(host_id) or 'N/A'
        bucket_name_dict["department"] = department_map.get(host_id, 'N/A')
        if not bucket_name_dict["scan_status"]:
            bucket_name_dict["scan_status"] = "N/A"
        if not bucket_name_dict['is_public']:
            bucket_name_dict['is_public'] = "Private"
        else:
            bucket_name_dict['is_public'] = "Public"
        if not bucket_name_dict['is_encrypted']:
            bucket_name_dict['is_encrypted'] = "Unencrypted"
        else :
            bucket_name_dict['is_encrypted'] = "Encrypted"
        del bucket_name_dict["id"]
        del bucket_name_dict["host_obj__id"]
    
    print(f'Creating DataFrame with {len(bucket_names_list)} rows')
    df = pd.DataFrame(bucket_names_list)
    print(f'DataFrame created with shape: {df.shape}')
    # 4) Rename columns to your desired field names
    df.rename(
        columns={
            "bucket_name": "Bucket Name",
            "region": "Region",
            "host_obj__scan_host": "Host",
            "host_tags": "Host Tags",
            "department": "Department",
            "is_public": "Public Access",
            "is_encrypted": "Encryption Status",
            "object_count": "Object Count",
            "scan_status": "Scan Status",
        },
        inplace=True
    )
    print(f'Renaming columns completed. DataFrame shape: {df.shape}')

    desired_order = [
        "Bucket Name", "Region", "Host", "Host Tags", "Department", "Public Access", "Encryption Status", "Object Count", "Scan Status"
    ]

    # This returns a new DataFrame with exactly that column sequence:
    df = df[desired_order]
    # serialize to pickle
    df.to_pickle(report_chunk)
    print(f'Chunked pickle written to: {report_chunk}')
    return report_chunk


@shared_task
def write_data_store_report_chunk(customer_db, report_id, chunked_ids, tmp_dir):
    """
    Outputs a pickle of a Pandas DataFrame with columns:
      Platform | Scanned host | Data Count | Data Types with data count | Data Type Subclassification
    """
    update_celery_config(customer_dbs_to_update=[customer_db])
    print("Dispatching chunk size:", len(chunked_ids))

    first_id = chunked_ids[0] if chunked_ids else "empty"
    last_id  = chunked_ids[-1] if chunked_ids else "empty"
    report_chunk = os.path.join(
        tmp_dir, f"chunked_data_store_table_{customer_db}_{report_id}_{first_id}-{last_id}.pkl"
    )
    print(f"report_chunk: {report_chunk}")

    cols = [
        "Platform",
        "Scanned host",
        "Data Count",
        "Data Types with data count",
        "Data Type Subclassification",
    ]

    # Handle empty chunk
    if not chunked_ids:
        os.makedirs(tmp_dir, exist_ok=True)
        pd.DataFrame(columns=cols).to_pickle(report_chunk)
        print(f"Chunked pickle written to: {report_chunk} (empty)")
        return report_chunk

    # Fetch scan breakdown
    scan_breakdown_list = list(
        ScanBreakdown.objects.using(customer_db)
        .filter(id__in=chunked_ids, data_count__gt=0)
        .annotate(
            platform=F("host_obj__scan_platform"),
            host_pk=F("host_obj__pk"),
            data_type=F("data_type_obj__data_type"),
        )
        .values("platform", "host_pk", "data_type", "data_count")
    )

    if not scan_breakdown_list:
        os.makedirs(tmp_dir, exist_ok=True)
        pd.DataFrame(columns=cols).to_pickle(report_chunk)
        print(f"Chunked pickle written to: {report_chunk} (empty)")
        return report_chunk

    df = pd.DataFrame(scan_breakdown_list)

    # Platform-wise totals
    platform_wise_data_count = (
        df.groupby("platform", dropna=False)["data_count"]
          .sum().reset_index().rename(columns={"data_count": "Data Count"})
    )
    platform_hosts = (
        df.groupby("platform", dropna=False)["host_pk"]
          .nunique().reset_index().rename(columns={"host_pk": "Scanned Host"})
    )

    # Data type-wise count per platform
    data_type_wise_count = (
        df.groupby(["platform", "data_type"], dropna=False)["data_count"]
          .sum().reset_index().rename(columns={"data_count": "dt_count"})
    )

    # Formatted data type string per platform
    formatted_data_type = (
        data_type_wise_count.sort_values(["platform", "dt_count", "data_type"], ascending=[True, False, True])
            .groupby("platform", dropna=False)
            .apply(lambda g: " | ".join(
                f"{t}:{int(c)}" for t, c in g[["data_type", "dt_count"]].values
                if isinstance(t, str) and t.strip()
            ) or "N/A")
            .reset_index(name="Data Types with data count")
    )

    # Subclassifications
    seen_data_types = [t for t in df["data_type"].dropna().astype(str).str.strip().unique() if t]
    subclassification_map = {}
    if seen_data_types:
        data_type_subclassification_row = list(
            DataType.objects.using(customer_db)
            .filter(data_type__in=seen_data_types)
            .annotate(subclassification=F("subclassification_objs__subclassification"))
            .values("data_type", "subclassification")
            .distinct()
        )
        for r in data_type_subclassification_row:
            dt = r["data_type"]
            sub = r["subclassification"]
            if not dt:
                continue
            subclassification_map.setdefault(dt, set())
            if isinstance(sub, str) and sub.strip():
                subclassification_map[dt].add(sub.strip())

    # Platform-wise subclassifications
    platform_subclassification_rows = []
    for plat, block in data_type_wise_count.groupby("platform", dropna=False):
        types_here = [t for t in block["data_type"] if isinstance(t, str) and t.strip()]
        subclassification = set().union(*(subclassification_map.get(t, set()) for t in types_here)) if types_here else set()
        platform_subclassification_rows.append({
            "platform": plat,  # lowercase internally
            "Data Type Subclassification": " | ".join(sorted(subclassification)) if subclassification else "N/A"
        })

    subclassification_df = pd.DataFrame(platform_subclassification_rows)

    # ----------------- Merge all -----------------
    final_dataframe = (platform_wise_data_count
           .merge(platform_hosts, on="platform", how="left")
           .merge(formatted_data_type, on="platform", how="left")
           .merge(subclassification_df, on="platform", how="left"))

    # Rename to match output columns
    final_dataframe = final_dataframe.rename(columns={"platform": "Platform"})

    # Fill missing / type cast
    final_dataframe["Data Types with data count"] = final_dataframe["Data Types with data count"].fillna("N/A")
    final_dataframe["Data Type Subclassification"] = final_dataframe["Data Type Subclassification"].fillna("N/A")
    final_dataframe["Data Count"] = final_dataframe["Data Count"].fillna(0).astype(int)
    final_dataframe["Scanned host"] = final_dataframe["Scanned host"].fillna(0).astype(int)

    # Reorder columns and sort
    final_dataframe = final_dataframe[cols].sort_values("Data Count", ascending=False, kind="stable")

    os.makedirs(tmp_dir, exist_ok=True)
    final_dataframe.to_pickle(report_chunk)
    print(f"Chunked pickle written to: {report_chunk}")
    return report_chunk


@shared_task
def write_asset_inventory_report_chunk(customer_db, report_id, chunked_ids, tmp_dir):
    """
    Outputs a pickle of a Pandas DataFrame with columns:
      Created At | Scan Host | Scan Platform | Host Tag | Department | Data Count | Sub-Classification
    """
    update_celery_config(customer_dbs_to_update=[customer_db])
    print("Dispatching chunk size:", len(chunked_ids))

    first_id = chunked_ids[0] if chunked_ids else "empty"
    last_id  = chunked_ids[-1] if chunked_ids else "empty"
    report_chunk = os.path.join(
        tmp_dir, f"chunked_host_table_{customer_db}_{report_id}_{first_id}-{last_id}.pkl"
    )
    print(f"report_chunk: {report_chunk}")

    cols = [
        "Created At",
        "Scan Host",
        "Scan Platform",
        "Host Tag",
        "Department",
        "Data Count",
        "Sub-Classification",
    ]

    os.makedirs(tmp_dir, exist_ok=True)

    if not chunked_ids:
        pd.DataFrame(columns=cols).to_pickle(report_chunk)
        print(f"Chunked pickle written to: {report_chunk} (empty)")
        return report_chunk

    # Query 1: Hosts, ordered by created_at (no Case/When)
    hosts = list(
        Host.objects.using(customer_db)
        .filter(pk__in=chunked_ids)
        .prefetch_related("host_tags", "department_objs")
        .only("id", "created_at", "scan_host", "scan_platform")
        .order_by("-created_at")
    )
    if not hosts:
        pd.DataFrame(columns=cols).to_pickle(report_chunk)
        print(f"Chunked pickle written to: {report_chunk} (no hosts)")
        return report_chunk

    host_ids = [h.id for h in hosts]

    # Query 2: Single ScanBreakdown fetch for BOTH totals and subclassifications
    subclassification_rows = list(
        ScanBreakdown.objects.using(customer_db)
        .filter(host_obj__id__in=host_ids, data_count__gt=0)
        .annotate(
            host_id=F("host_obj__id"),
            sb_id=F("id"),
            subclassification=F("data_type_obj__subclassification_objs__subclassification"),
        )
        .values("host_id", "sb_id", "data_count", "subclassification")
    )

    if subclassification_rows:
        sb_df = pd.DataFrame(subclassification_rows)

        # Sum data_count per host WITHOUT double-counting: one row per sb_id
        totals = (
            sb_df.drop_duplicates("sb_id")
                 .groupby("host_id", dropna=False)["data_count"]
                 .sum()
                 .astype(int)
        )
        data_count_map = totals.to_dict()

        # Collect distinct subclassifications per host from same rows
        subclassification_map = (
            sb_df.dropna(subset=["subclassification"])
                 .assign(subclassification=sb_df["subclassification"].astype(str).str.strip())
        )
        subclassification_map = (
            subclassification_map[subclassification_map["subclassification"] != ""]
            .groupby("host_id", dropna=False)["subclassification"]
            .apply(lambda s: " | ".join(sorted(set(s))))
            .to_dict()
        )
    else:
        data_count_map, subclassification_map = {}, {}

    # Build rows
    rows = []
    for h in hosts:
        created = None
        if h.created_at:
            created = (
                h.created_at.strftime(r"%b. %d, %Y, %I:%M %p")
                .lstrip("0")
                .replace(" 0", " ")
            )

        host_tags_str = " | ".join(sorted({t.tag_name.strip() for t in h.host_tags.all() if t.tag_name})) or "N/A"
        dept_str = " | ".join(sorted({d.department_name.strip() for d in h.department_objs.all() if d.department_name})) or "N/A"
        subs_str = subclassification_map.get(h.id, "N/A")

        rows.append({
            "Created At": created,
            "Scan Host": h.scan_host,
            "Scan Platform": h.scan_platform,
            "Host Tag": host_tags_str,
            "Department": dept_str,
            "Data Count": int(data_count_map.get(h.id, 0)),
            "Sub-Classification": subs_str,
        })

    df = pd.DataFrame(rows, columns=cols)
    df.to_pickle(report_chunk)
    print(f"Chunked pickle written to: {report_chunk}")
    return report_chunk



@shared_task
def write_scan_breakdown_report_chunk(customer_db, report_id, chunked_ids, tmp_dir):
    """
    Build a DataFrame with columns:
      Host | Scan Platform | Host Tag | Department | Scan ID | Data Location | Data Type |
      Sub-Classification | Data Count | Data Extracted
    and write it as a pickle to tmp_dir.
    """
    # Best-effort celery config update
    try:
        update_celery_config(customer_dbs_to_update=[customer_db])
    except Exception:
        pass

    # Chunk file path (safe on empty)
    first_id = chunked_ids[0] if chunked_ids else "none"
    last_id  = chunked_ids[-1] if chunked_ids else "none"
    out_path = os.path.join(
        tmp_dir, f"scan_breakdown_report_chunk_{customer_db}_{report_id}_{first_id}-{last_id}.pkl"
    )
    os.makedirs(tmp_dir, exist_ok=True)
    print(f"[write_scan_breakdown_report_chunk] writing: {out_path} (count={len(chunked_ids)})")

    # Empty guard
    columns = [
        "Host", "Scan Platform", "Host Tag", "Department", "Scan ID",
        "Data Location", "Data Type", "Sub-Classification", "Data Count", "Data Extracted"
    ]
    if not chunked_ids:
        pd.DataFrame(columns=columns).to_pickle(out_path)
        print(f"[write_scan_breakdown_report_chunk] empty chunk written")
        return out_path

    qs = (
        ScanBreakdown.objects.using(customer_db)
        .filter(pk__in=chunked_ids)
        .select_related(
            "scan_details_obj",
            "scan_details_obj__scan_unit_obj",
            "scan_details_obj__scan_obj",
            "scan_details_obj__scan_obj__host_obj",
            "data_type_obj",
        )
        .prefetch_related(
            "scan_details_obj__scan_obj__host_obj__host_tags",
            "scan_details_obj__scan_obj__host_obj__department_objs",
            "data_type_obj__subclassification_objs",
        )
        .order_by("-id")
    )

    rows = []

    for scan_breakdown_obj in qs:
        scan_detail_obj  = scan_breakdown_obj.scan_details_obj
        scan_unit_obj  = scan_detail_obj.scan_unit_obj if scan_detail_obj else None
        scan_object  = scan_detail_obj.scan_obj if scan_detail_obj else None
        host_object  = scan_object.host_obj if scan_object else None
        data_type_object = scan_breakdown_obj.data_type_obj

        host_name      = getattr(host_object, "scan_host", "") if host_object else ""
        scan_platform  = getattr(host_object, "scan_platform", "") if host_object else ""
        scan_id        = getattr(scan_object, "id", None)

        # Host Tag / Department (pipe-joined, unique)
        host_tags_str = " | ".join(sorted({t.tag_name.strip() for t in (host_object.host_tags.all() if host_object else []) if t.tag_name})) or "N/A"
        dept_str      = " | ".join(sorted({d.department_name.strip() for d in (host_object.department_objs.all() if host_object else []) if d.department_name})) or "N/A"

        # Data Type
        data_type = getattr(data_type_object, "data_type", "") if data_type_object else ""

        # Sub-Classification (pipe-joined unique across the data_type's subclassification set)
        sub_str = " | ".join(sorted({
            sc.subclassification.strip()
            for sc in (data_type_object.subclassification_objs.all() if data_type_object else [])
            if isinstance(sc.subclassification, str) and sc.subclassification.strip()
        })) or "N/A"

        # Data Count and Extracted flag
        data_count = int(scan_breakdown_obj.data_count or 0)
        extracted  = "Yes" if bool(getattr(scan_breakdown_obj, "data_extracted", False)) else "No"

        # Data Location (simple, readable pick per platform type)
        data_loc = ""
        platform_type = getattr(host_object, "platform_type", None) if host_object else None
        

        if platform_type in ['Drive', 'Cloud']:
            data_loc_1 = getattr(scan_unit_obj, "file_path", None) or getattr(scan_unit_obj, "file_name", None) or ""
            data_loc = f"File Path: {data_loc_1}" if data_loc_1 else "N/A"
        elif platform_type == 'DB':
            db   = getattr(scan_unit_obj, "db_name", "") or ""
            tbl  = getattr(scan_unit_obj, "table_name", "") or ""
            col  = getattr(scan_unit_obj, "column_name", "") or ""
            data_loc = f"DB Name: {db}, Table: {tbl}, Column: {col}" if db or tbl or col else "N/A"
        elif platform_type == 'Email':
            subj = getattr(scan_unit_obj, "email_subject", "") or ""
            data_loc = f"Subject: {subj}" if subj else "N/A"
        elif platform_type == 'App':
            if scan_platform == 'Jira':
                proj = getattr(scan_unit_obj, "project_name", "") or ""
                iid  = getattr(scan_unit_obj, "issue_id", "") or ""
                data_loc = f"Project: {proj}"
            else:
                cname = getattr(scan_unit_obj, "chat_name", "") or ""
                cid   = getattr(scan_unit_obj, "chat_id", "") or ""
                data_loc = f"{cname} ({cid})" if cname or cid else ""
        # fallback to unit id if nothing else
        if not data_loc:
            data_loc = str(getattr(scan_unit_obj, "id", "") or "")

        rows.append({
            "Host": host_name or "N/A",
            "Scan Platform": scan_platform or "N/A",
            "Host Tag": host_tags_str,
            "Department": dept_str,
            "Scan ID": scan_id if scan_id is not None else "",
            "Data Location": data_loc or "N/A",
            "Data Type": data_type or "N/A",
            "Sub-Classification": sub_str,
            "Data Count": data_count,
            "Data Extracted": extracted,
        })

    df = pd.DataFrame(rows, columns=columns)
    df.to_pickle(out_path)
    print(f"[write_scan_breakdown_report_chunk] chunk written: {out_path} rows={len(df)}")
    return out_path



def gen_report(customer_db: str, report_id: int):
    """
    Generates different type of report.
    """

    report_obj_up = GenReport.objects.using(customer_db).filter(id=int(report_id))
    report_obj = report_obj_up.first()
    report_id = report_obj.id
    report_type = report_obj.report_type
    report_format = report_obj.report_format
    worker_limit = report_obj.worker_limit
    filter_inputs = report_obj.filter_inputs

    report_type = str(report_type).lower().strip()
    report_extension = "xlsx" if report_format.lower() == "excel" else "csv"
    report_name = f"{report_type}_{report_id}.{report_extension}"
    report_path = os.path.join(settings.MEDIA_ROOT, report_name)

    # Getting connector object.
    connector_obj = get_connector_obj(
        connector_type=CONNECTOR_TYPE,
        private_ip=PRIVATE_IP,
        customer_db=customer_db,
    )

    report_obj_up.update(
        connector_obj=connector_obj,
        report_name=report_name, 
        report_status="Processing"
    )

    if report_type == "scan_report":
        filtered_result_objs = filter_result_objs(customer_db, filter_inputs)
        qs = filtered_result_objs
        write_report_chunk = write_scan_report_chunk
        sheet_title = "Scan Report"
    
    elif report_type == "scan_status_report":
        filtered_scan_objs = filter_scan_objs(customer_db, filter_inputs)
        # filter_inputs["selected_scan_objs"] = filtered_scan_objs
        # filtered_scan_breakdown_objs = filter_scan_breakdown_objs(customer_db, filter_inputs)
        qs = filtered_scan_objs
        write_report_chunk = write_scan_status_report_chunk
        sheet_title = "Scan Status Report"

    elif report_type == "data_report":
        scan_breakdown_objs = filter_scan_breakdown_objs(customer_db, filter_inputs)
        filter_inputs["selected_scan_breakdown_objs"] = scan_breakdown_objs
        filtered_result_objs = filter_result_objs(customer_db, filter_inputs)
        data_ids = filtered_result_objs.values_list('data_obj__id', flat=True).distinct()
        filtered_data_objs = Data.objects.using(customer_db).filter(id__in=data_ids)
        qs = filtered_result_objs
        print(f'filtered_data_objs: {filtered_data_objs.count()}')
        write_report_chunk = write_data_report_chunk
        sheet_title = "Data Report"

    elif report_type == "db_report":
        filtered_db_name_objs = filter_db_name_objs(customer_db, filter_inputs)
        qs = filtered_db_name_objs
        write_report_chunk = write_db_report_chunk
        sheet_title = "DB Report"

    elif report_type == "table_report":
        filtered_table_name_objs = filter_table_name_objs(customer_db, filter_inputs)
        qs = filtered_table_name_objs
        write_report_chunk = write_table_report_chunk
        sheet_title = "Table Report"
    
    elif report_type == "bucket_name_report":
        filtered_bucket_name_objs = filter_bucket_name_objs(customer_db, filter_inputs)
        qs = filtered_bucket_name_objs
        write_report_chunk = write_bucket_name_report_chunk
        sheet_title = "Bucket Name Report"

    elif report_type == "scan_unit_report":
        filtered_scan_unit_objs = filter_scan_unit_objs(customer_db, filter_inputs)
        qs = filtered_scan_unit_objs
        write_report_chunk = write_scan_unit_report_chunk
        sheet_title = "Scan Unit Report"
    
    elif report_type == "scan_details_report":
        filtered_column_name_objs = filter_scan_details_objs(customer_db, filter_inputs)
        qs = filtered_column_name_objs
        write_report_chunk = write_scan_details_report_chunk
        sheet_title = "Scan Details Report"

    elif report_type == "column_report":
        filtered_column_name_objs = filter_column_name_objs(customer_db, filter_inputs)
        qs = filtered_column_name_objs
        write_report_chunk = write_column_report_chunk
        sheet_title = "Column Report"

    elif report_type == "host_details_report":
        filtered_host_objs = filter_host_objs(customer_db, filter_inputs)
        qs = filtered_host_objs
        write_report_chunk = write_host_details_report_chunk
        sheet_title = "Host Details Report"
    
    elif report_type == "audit_log_report":
        filtered_audit_log_objs = filter_audit_log_objs(customer_db, filter_inputs)
        qs = filtered_audit_log_objs
        write_report_chunk = write_audit_log_report_chunk
        sheet_title = "Audit Log Report"

    elif report_type == "found_data_type_report":
        scan_breakdown_objs = filter_scan_breakdown_objs(customer_db, filter_inputs).filter(
            data_count__gt=0
        )
        data_type_ids = scan_breakdown_objs.values_list("data_type_obj", flat=True).distinct()
        filtered_data_type_objs = filter_data_type_objs(customer_db, filter_inputs).filter(id__in=data_type_ids)
        qs = filtered_data_type_objs
        write_report_chunk = write_found_data_type_report_chunk
        sheet_title = "Found Data Type Report"

    elif report_type == "host_user_detail_report":
        filtered_host_user_objs = filter_host_user_objs(customer_db, filter_inputs)
        print(f"filtered_host_user_objs: {filtered_host_user_objs.count()}")
        qs = filtered_host_user_objs
        write_report_chunk = write_host_user_detail_report_chunk
        sheet_title = "Host User Detail Report"

    elif report_type == "data_store_report":
        filtered_scan_breakdown_objs = filter_scan_breakdown_objs(customer_db, filter_inputs)
        filtered_data_type_objs = filter_data_type_objs(customer_db, filter_inputs)
        qs = filtered_scan_breakdown_objs
        write_report_chunk = write_data_store_report_chunk
        sheet_title = "Data Store Report"

    elif report_type == "asset_inventory_report":
        filtered_host_objs = filter_host_objs(customer_db, filter_inputs)
        qs = filtered_host_objs
        write_report_chunk = write_asset_inventory_report_chunk
        sheet_title = "Asset Inventory Report"

    elif report_type == "scan_breakdown_report":
        filtered_scan_breakdown_objs = filter_scan_breakdown_objs(customer_db, filter_inputs)
        qs = filtered_scan_breakdown_objs
        write_report_chunk = write_scan_breakdown_report_chunk
        sheet_title = "Asset Inventory Report"

    total_items = qs.count()
    chunk_size = settings.REPORT_CHUNK_SIZE
    progress_per_chunk = 100 / (total_items // chunk_size + 1) if total_items > chunk_size else 100

    report_obj_up.update(report_status="Writing")

    assigned_tasks = []
    report_chunks = []
    progress = 0
    for chunked_ids in get_chunked_ids(queryset=qs, chunk_size=chunk_size):
        task = write_report_chunk.delay(
            customer_db=customer_db,
            report_id=report_id,
            chunked_ids=chunked_ids,
            tmp_dir=settings.MEDIA_ROOT,
        )
        assigned_tasks.append(task)

        while len(assigned_tasks) >= worker_limit:
            for task in assigned_tasks.copy():
                try:
                    if task.ready():
                        report_chunks.append(task.get())
                        assigned_tasks.remove(task)
                        progress += progress_per_chunk
                        print(f"writing progress: {progress}%")
                except Exception as e:
                    print(f"Error processing task {task.id}: {e}")
                    assigned_tasks.remove(task)
                    progress += progress_per_chunk
            
            report_progress = round(progress, 2)
            report_obj_up.update(report_progress=report_progress)
            time.sleep(1)

    # Handling any remaining tasks.
    while len(assigned_tasks) > 0:
        for task in assigned_tasks.copy():
            try:
                if task.ready():
                    report_chunks.append(task.get())
                    assigned_tasks.remove(task)
                    progress += progress_per_chunk
                    print(f"writing progress: {progress}%")
            except Exception as e:
                print(f"Error processing task {task.id}: {e}")
                assigned_tasks.remove(task)
                progress += progress_per_chunk
        report_progress = round(progress, 2)
        report_obj_up.update(report_progress=report_progress)
        time.sleep(1)

    
    report_obj_up.update(report_status="Merging")
    # Merging all chunked reports into a single file.
    merge_report_chunks(
        customer_db=customer_db,
        report_id=report_id,
        report_chunks=report_chunks,
        merged_report_path=report_path,
        sheet_title=sheet_title
    )

    # Final updating the report object.
    report_obj_up.update(
        report_path=report_path,
        report_status="Generated",
    )

    print(f"Report ({report_type}) generated: {report_path}")
    return

